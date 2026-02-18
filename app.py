import os
import sqlite3
import unicodedata
from pathlib import Path
from typing import Optional, Tuple, List

from fastapi import FastAPI, Query, HTTPException, UploadFile, File, Request
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles

import openpyxl

# =========================
# CONFIG
# =========================
APP_DIR = Path(__file__).resolve().parent
templates = Jinja2Templates(directory=str(APP_DIR / "templates"))

STORAGE_DIR = Path(os.getenv("TDFC_STORAGE_DIR", APP_DIR / "storage"))
STORAGE_DIR.mkdir(parents=True, exist_ok=True)

DATA_FILE = STORAGE_DIR / "current.xlsx"
DB_PATH = STORAGE_DIR / "tdfc_cache.sqlite"

SHEET_NAME_DEFAULT = os.getenv("TDFC_SHEET", "2026")
ADMIN_KEY = os.getenv("TDFC_ADMIN_KEY", "")  # recommandé

# =========================
# APP
# =========================
app = FastAPI(title="TDFC Dico", version="4.0")

# Static (ne plante pas si le dossier n'existe pas)
STATIC_DIR = APP_DIR / "static"
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


# =========================
# UTILS
# =========================
def norm(s) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = "".join(ch for ch in unicodedata.normalize("NFD", s) if unicodedata.category(ch) != "Mn")
    s = s.replace("\u00a0", " ")
    while "  " in s:
        s = s.replace("  ", " ")
    return s


def require_admin(key: str) -> None:
    # Si ADMIN_KEY est vide => pas de protection (déconseillé)
    if ADMIN_KEY and key != ADMIN_KEY:
        raise HTTPException(status_code=401, detail="Unauthorized")


def ensure_db(conn: sqlite3.Connection) -> None:
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    conn.execute("PRAGMA temp_store=MEMORY;")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS meta (
          key TEXT PRIMARY KEY,
          value TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS entries (
          imprime TEXT NOT NULL,
          codeedi TEXT NOT NULL,
          libelle TEXT NOT NULL
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_entries ON entries(imprime, codeedi)")
    conn.commit()


def get_file_sig(sheet_name: str) -> str:
    if not DATA_FILE.exists():
        return ""
    st = DATA_FILE.stat()
    return f"{DATA_FILE.resolve()}|{int(st.st_mtime)}|{st.st_size}|{sheet_name}"


def find_header(ws, max_scan_rows: int = 40) -> Tuple[int, int, int, int]:
    def pick(hm, *keys) -> Optional[int]:
        for k in keys:
            nk = norm(k)
            if nk in hm:
                return hm[nk]
        return None

    max_r = min(max_scan_rows, ws.max_row or 0)
    for r in range(1, max_r + 1):
        hm = {}
        for idx, cell in enumerate(ws[r], start=1):
            t = norm(cell.value)
            if t and t not in hm:
                hm[t] = idx

        ci = pick(hm, "imprimé", "imprime")
        cc = pick(hm, "code edi", "code_edi", "codeedi")
        cl = pick(hm, "libellé", "libelle")

        if ci and cc and cl:
            return r, ci, cc, cl

    raise RuntimeError("En-têtes introuvables (Imprimé / Code EDI / Libellé).")


def rebuild_cache(sheet_name: str) -> None:
    if not DATA_FILE.exists():
        raise RuntimeError("Aucun fichier uploadé. Uploade d'abord un .xlsx.")

    with sqlite3.connect(DB_PATH) as conn:
        ensure_db(conn)
        file_sig = get_file_sig(sheet_name)

        conn.execute("DELETE FROM entries;")
        conn.execute("DELETE FROM meta;")
        conn.execute("INSERT INTO meta(key, value) VALUES(?, ?)", ("file_sig", file_sig))
        conn.commit()

        wb = openpyxl.load_workbook(DATA_FILE, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"Onglet '{sheet_name}' introuvable.")
        ws = wb[sheet_name]

        header_row, col_imprime, col_codeedi, col_libelle = find_header(ws)

        batch = []
        BATCH_SIZE = 3000

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if col_imprime - 1 >= len(row) or col_codeedi - 1 >= len(row):
                continue

            imp = norm(row[col_imprime - 1])
            edi = norm(row[col_codeedi - 1])
            if not imp or not edi:
                continue

            lib = ""
            if col_libelle - 1 < len(row) and row[col_libelle - 1] is not None:
                lib = str(row[col_libelle - 1]).strip()
            if not lib:
                continue

            batch.append((imp, edi, lib))
            if len(batch) >= BATCH_SIZE:
                conn.executemany("INSERT INTO entries(imprime, codeedi, libelle) VALUES(?, ?, ?)", batch)
                conn.commit()
                batch.clear()

        if batch:
            conn.executemany("INSERT INTO entries(imprime, codeedi, libelle) VALUES(?, ?, ?)", batch)
            conn.commit()


def ensure_cache_uptodate(sheet_name: str) -> None:
    if not DATA_FILE.exists():
        raise RuntimeError("Aucun fichier uploadé. Ouvre /?admin=1 pour uploader le .xlsx.")

    with sqlite3.connect(DB_PATH) as conn:
        ensure_db(conn)
        cur = conn.cursor()
        cur.execute("SELECT value FROM meta WHERE key='file_sig'")
        row = cur.fetchone()
        cached_sig = row[0] if row else ""
        current_sig = get_file_sig(sheet_name)

        if cached_sig != current_sig:
            rebuild_cache(sheet_name)


# =========================
# ROUTES
# =========================
@app.get("/", response_class=HTMLResponse)
def home(request: Request, admin: int = Query(0)):
    has_file = DATA_FILE.exists()
    is_admin_ui = (admin == 1)

    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "has_file": has_file,
            "sheet_default": SHEET_NAME_DEFAULT,
            "is_admin_ui": is_admin_ui,
        },
    )


@app.post("/upload")
async def upload_excel(
    file: UploadFile = File(...),
    sheet: str = Query(SHEET_NAME_DEFAULT),
    key: str = Query(""),
):
    require_admin(key)

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Format invalide. Upload uniquement .xlsx")

    content = await file.read()
    if len(content) < 1000:
        raise HTTPException(status_code=400, detail="Fichier invalide ou trop petit")

    tmp_path = STORAGE_DIR / "upload.tmp.xlsx"
    tmp_path.write_bytes(content)
    tmp_path.replace(DATA_FILE)

    try:
        rebuild_cache(sheet)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    return {"ok": True, "message": "Fichier uploadé et indexé", "sheet": sheet}


@app.get("/lookup")
def lookup(
    imprime: str = Query(..., min_length=1),
    codeedi: str = Query(..., min_length=1),
    all: bool = False,
    sheet: str = Query(SHEET_NAME_DEFAULT),
):
    try:
        ensure_cache_uptodate(sheet)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    imp = norm(imprime)
    edi = norm(codeedi)

    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()

        if not all:
            cur.execute(
                "SELECT libelle FROM entries WHERE imprime=? AND codeedi=? LIMIT 1",
                (imp, edi),
            )
            r = cur.fetchone()
            return {"found": bool(r), "libelle": (r[0] if r else "")}

        cur.execute("SELECT libelle FROM entries WHERE imprime=? AND codeedi=?", (imp, edi))
        rows = [x[0] for x in cur.fetchall()]

        seen = set()
        out: List[str] = []
        for x in rows:
            k = norm(x)
            if k not in seen:
                seen.add(k)
                out.append(x)

        return {"found": len(out) > 0, "libelles": out}
